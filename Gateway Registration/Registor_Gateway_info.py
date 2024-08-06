'''
    Automation for GStack project
    this script will automate the process of creating a new Gateway to GStack, https://gstack.gwin.emsd.gov.hk/
    1) Login to GStack
    2) Go to Gateway page
    3) Create a new Gateway in the Machine only page
'''

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
from typing import List
import os

class GStack_Webpage:
    def __init__(self):
        self.Domain = 'https://gstack.gwin.emsd.gov.hk/'
        self.User_Name = [LOGIN_NAME] # <-- Login user name
        self.Password = [PASSWORD] # <-- Login Password


class EMSD_Project:
    def __init__(self):
        self.organization = "EMSD"
        self.project_code_SLP = [PROJECT_01] # <-- Project code 01
        self.project_code_FSD = [PROJECT_02] # <-- Project code 02
        self.project_code_DSD = [PROJECT_03] # <-- Project code 03
        self.project_Owner = [OWNER] # <-- Owner of this project

class GatewayInfo:
    def __init__(self, in_gateway_EUI, in_project_Code, in_owner, in_gateway_brand, in_gateway_model, in_antenna_type, in_antenna_dbi, in_Gateway_Firmware_version, in_VPN_KEY, in_SSH_USER, in_SSH_PASSWORD, in_Default_USER, in_Default_PASSWORD, in_Serial, in_Token):
        self.EUI = in_gateway_EUI.lower() if in_gateway_EUI is not None else ""
        self.Project_Code = in_project_Code
        self.Owner = in_owner
        self.Brand = in_gateway_brand
        self.Model = in_gateway_model
        self.Antenna_Type = in_antenna_type
        self.Antenna_DBI = in_antenna_dbi
        self.Firmware_Version = in_Gateway_Firmware_version
        self.VPN_KEY = in_VPN_KEY
        self.SSH_USER = in_SSH_USER
        self.SSH_PASSWORD = in_SSH_PASSWORD
        self.Default_USER = in_Default_USER
        self.Default_PASSWORD = in_Default_PASSWORD
        self.SERIAL = in_Serial
        self.Gateway_Token = in_Token

class Excel_Controller:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def read_excel_and_convert_to_gateway_info(self, page_name) -> List[GatewayInfo]:
        workbook = openpyxl.load_workbook(self.file_path)
        result_sheet = workbook[page_name]
        gateway_info_list = []

        for row in result_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            gateway_info = GatewayInfo(*row)
            gateway_info_list.append(gateway_info)

        return gateway_info_list

def login(driver, domain_login, _user_name, _password):
    driver.get(domain_login)
    driver.find_element(By.NAME, 'username').send_keys(_user_name)
    time.sleep(1)
    driver.find_element(By.NAME, 'password').send_keys(_password)
    time.sleep(1)

    driver.find_element(By.ID, 'login-button').click()
    # wait the ready state to be complete
    WebDriverWait(driver=driver, timeout=10).until(
        lambda x: x.execute_script("return document.readyState === 'complete'")
    )
    error_message = "Incorrect username or password."
    errors = driver.find_elements("css selector", ".flash-error")
    if any(error_message in e.text for e in errors):
        print("[!] Login failed")
    else:
        print("[+] Login successful")

    return driver
    

def goto_Gateway_Page(driver, sleep_time):
    # Go to Gateway page
    driver.find_element(By.ID, 'nav-bar-button').click()
    time.sleep(sleep_time)
    driver.find_element(By.ID, 'left-menu-gateway').click()
    time.sleep(sleep_time*2)
    return driver

def goto_Add_Gateway_Page(driver):
    # Go to Add Gateway page
    driver.find_element(By.CLASS_NAME, 'page-add-button').click()
    time.sleep(2)
    return driver

def goto_hardware_page(driver, sleep_time):
    # Go to Hardware page
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div/div[1]/a[5]').click()
    time.sleep(sleep_time)
    return driver

def setup_Organization(driver, in_organization, sleep_time):
    # Select Organization Dropdown
    # Active the dropdown option --> //*[@id="organization-select-dropdown"]/input
    # //*[@id="organization-select-dropdown"]/i
    driver.find_element(By.XPATH, '//*[@id="organization-select-dropdown"]/i').click()
    time.sleep(sleep_time/2)
    # if EMSD --> '//*[@id="organization-select-dropdown"]/div[2]/div[1]' click
    if(in_organization == "EMSD"):
        driver.find_element(By.XPATH, '//*[@id="organization-select-dropdown"]/div[2]/div[2]').click()
        time.sleep(sleep_time)
    return driver

def setup_Project_Department(driver, project_Code, time_sleep):
    
    # Select
    project = EMSD_Project()
    if(project.organization =="EMSD"):
        driver.find_element(By.XPATH, '//*[@id="project-select-dropdown"]').click()
        time.sleep(time_sleep*2)
        # Select Project Dropdown
        if(project_Code == project.project_code_SLP):
            driver.find_element(By.XPATH, '//*[@id="project-select-dropdown"]/div/div[1]').click()
        elif(project_Code == project.project_code_FSD):
            driver.find_element(By.XPATH, '//*[@id="project-select-dropdown"]/div/div[2]').click()
        elif(project_Code == project.project_code_DSD):
            driver.find_element(By.XPATH, '//*[@id="project-select-dropdown"]/div/div[3]').click()
        driver.find_element(By.TAG_NAME, 'form').submit()
        time.sleep(time_sleep/4)

        driver.find_element(By.XPATH, '//*[@id="toggle_HARDWARE_Btn"]').click()
        time.sleep(time_sleep*2)

    return driver

def setup_Gateway_Info(driver, in_gateway_info: GatewayInfo, time_sleep):
    # Input Gateway informations
    # Gateway ID --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[4]/div[1]/div[1]/input'
    # //*[@id="segment-box"]/div/div[2]/form[2]/div[4]/div[1]/div[1]/input
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[4]/div[1]/div[1]/input').clear()
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[4]/div[1]/div[1]/input').send_keys(in_gateway_info.EUI)
    print(in_gateway_info.EUI + " is Added")
    time.sleep(time_sleep/4)
    
    # Contract No. --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[5]/div[1]/div/input'
    # //*[@id="segment-box"]/div/div[2]/form[2]/div[5]/div[1]/div/input
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[5]/div[1]/div/input').clear()
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[5]/div[1]/div/input').send_keys(in_gateway_info.Project_Code)
    print(in_gateway_info.Project_Code + " is Added")
    time.sleep(time_sleep/4)

    # Owner -->'/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[5]/div[2]/div/input'
    # //*[@id="segment-box"]/div/div[2]/form[2]/div[5]/div[2]/div/input
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[5]/div[2]/div/input').clear()
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[5]/div[2]/div/input').send_keys(in_gateway_info.Owner)
    print(in_gateway_info.Owner + " is Added")
    time.sleep(time_sleep/4)
    
    # Brand --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[1]/div/input'
    #driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[1]/div/input').send_keys(in_gateway_info.Brand)
    driver = set_Brand(driver, in_gateway_info.Brand, time_sleep)
    print(in_gateway_info.Brand + " is Added")
    time.sleep(time_sleep/4)

    
    # Model --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[2]/div/input'
    #driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[2]/div/input').send_keys(in_gateway_info.Model) 
    driver = set_model(driver, in_gateway_info.Model, in_gateway_info.Brand, time_sleep)
    print(in_gateway_info.Model + " is Added")   
    time.sleep(time_sleep/4)

    # Antenna Type --> '#Antenna Type --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[2]/div/input'
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[1]/div/input').send_keys(in_gateway_info.Antenna_Type)
    print(in_gateway_info.Antenna_Type + " is Added")
    time.sleep(time_sleep/4)

    driver = set_antenna_dbi(driver, in_gateway_info.Antenna_DBI, time_sleep)
    print(str(in_gateway_info.Antenna_DBI) + " is Added")
    time.sleep(time_sleep/4)

    # Firmware version --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[3]/div/input'
    driver.find_element(By.NAME, 'hardware.firmware_version').send_keys(in_gateway_info.Firmware_Version) 
    print(in_gateway_info.Firmware_Version + " is Added")
    time.sleep(time_sleep/4)

    # OpenVPN Cert Name --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[4]/div/input'
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[4]/div/input').send_keys(in_gateway_info.VPN_KEY)
    print(in_gateway_info.VPN_KEY + " is Added")
    time.sleep(time_sleep/4)

    # SSH User name --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[1]/div/input'
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[1]/div/input').send_keys(in_gateway_info.SSH_USER)
    print(in_gateway_info.SSH_USER + " is Added")
    time.sleep(time_sleep/4)

    # SSH Password --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[2]/div/input'
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[2]/div/input').send_keys(in_gateway_info.SSH_PASSWORD)
    print(in_gateway_info.SSH_PASSWORD + " is Added")
    time.sleep(time_sleep/4)

    # Default User Name -->'/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[3]/div/input'
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[3]/div/input').send_keys(in_gateway_info.Default_USER) 
    print(in_gateway_info.Default_USER + " is Added")
    time.sleep(time_sleep/4)

    # Default Password --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[4]/div/input'
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[4]/div/input').send_keys(in_gateway_info.Default_PASSWORD)
    print(in_gateway_info.Default_PASSWORD + " is Added")
    time.sleep(time_sleep/4)

    # Click on the Save button
    driver.find_element(By.ID, 'gateway-hardware-form-button').click()
    time.sleep(time_sleep*8)
    
    # Click on the Submit button
    driver.find_element(By.ID, 'gateway-hardware-form-submit-button').click()
    time.sleep(time_sleep*8)

    # Click OK button for Record success
    driver.find_element(By.XPATH, '//*[@id="popUp-message-container"]/div[3]/button').click()  
    time.sleep(sleep_time/4)
    return driver

def set_model(driver , in_model, in_brand, time_sleep):
    # 1) Select Model Dropdown
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[6]/div[2]/div/i').click()
    time.sleep(time_sleep/4)
    # Model --> RAK7289CV2
    # /html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[2]/div/div[2]/div[5]
    # Model --> RAK7268CV2
    # /html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[2]/div/div[2]/div[6]
    brand = in_brand.lower()
    model = in_model.lower()
    if(brand =="rak"):
        if(model == "rak7268"):
            driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[6]/div[2]/div/div[2]/div[6]').click()
        elif(model == "rak7289"):
            driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[6]/div[2]/div/div[2]/div[5]').click()
    time.sleep(time_sleep/4)
    return driver

def set_Brand(driver, in_brand, time_sleep):
    #1) Select Brand Dropdown
    # //*[@id="segment-box"]/div/div[2]/form[2]/div[6]/div[1]/div/i
    driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[6]/div[1]/div/i').click()
    time.sleep(time_sleep/4)
    brand = in_brand.lower()
    #2) Select Brand: RA
    # rak --> '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[6]/div[1]/div/div[2]/div[3]' click
    if(brand == "rak"):
        driver.find_element(By.XPATH, '//*[@id="segment-box"]/div/div[2]/form[2]/div[6]/div[1]/div/div[2]/div[3]').click()
    time.sleep(time_sleep/4)
    return driver


def set_antenna_dbi(driver, in_antenna_dbi, time_sleep):
    #Set Antenna Power
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[2]/div/i').click()
    time.sleep(time_sleep)
    if(in_antenna_dbi == 2.6):
        driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[2]/div/div[2]/div[1]').click()
    elif(in_antenna_dbi == 3):
        driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[2]/div/div[2]/div[2]').click()
    elif(in_antenna_dbi == 6):
        driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[2]/div/div[2]/div[3]').click()
    elif(in_antenna_dbi == 9):
        driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[7]/div[2]/div/div[2]/div[4]').click()
    time.sleep(time_sleep)
    return driver

def get_Gateway_Token(driver, in_gateway_info: GatewayInfo, sleep_time):
    time.sleep(sleep_time*5)
    # Gateway ID = in_gateway_info.EUI
    # //*[@id="homepage-second-row"]/div/div/div[2]/div/div[1]/div[2]/div/div[4]/input send_keys
    driver.find_element(By.XPATH, "//*[@id='homepage-second-row']/div/div/div[2]/div/div[1]/div[2]/div/div[4]/input").send_keys(in_gateway_info.EUI)
    # //*[@id="homepage-second-row"]/div/div/div[2]/div/div[1]/div[2]/div/div[1]/button click
    driver.find_element(By.XPATH, "//*[@id='homepage-second-row']/div/div/div[2]/div/div[1]/div[2]/div/div[1]/button").click()
    time.sleep(sleep_time*2)

    # Click on the Gateway EUI
    driver.find_element(By.XPATH, f"//a[@href='/gateway-hardware-details?dev_eui={in_gateway_info.EUI.lower()}']").click()
    time.sleep(5)

    # Switch to the new window
    new_window = driver.window_handles[-1]
    driver.switch_to.window(new_window)

    # Wait for the input element to be loaded
    wait = WebDriverWait(driver, 20)  # Wait up to 15 seconds
    token_result_element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div/div[2]/div/div[2]/div/div/div/div[2]/form[2]/div[8]/div[1]/div/input")))

    token_result = token_result_element.get_attribute("value")
    in_gateway_info.Gateway_Token = token_result

    return driver, in_gateway_info


#read from excel file for Gateway informations
def read_Gateway_info_from_Excel(in_file_path: str, page_name):
    excel_controller = Excel_Controller(in_file_path)
    gatewayInfos = excel_controller.read_excel_and_convert_to_gateway_info(page_name)
    return gatewayInfos

def save_gateway_info_to_excel(gateway_info_list: List[GatewayInfo], _finished_data_file_path: str, _sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name

    # Write header
    headers = ['EUI', 
               'Project_Code', 
               'Owner', 
               'Brand', 
               'Model', 
               'Antenna_Type', 
               'Antenna_DBI', 
               'Firmware_Version', 
               'VPN_KEY', 
               'SSH_USER', 
               'SSH_PASSWORD', 
               'Default_USER', 
               'Default_PASSWORD', 
               'SERIAL',
               'Gateway_Token']
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i, value=header)

    # Write data
    for i, gateway_info in enumerate(gateway_info_list, start=2):
        sheet.cell(row=i, column=1, value=gateway_info.EUI)
        sheet.cell(row=i, column=2, value=gateway_info.Project_Code)
        sheet.cell(row=i, column=3, value=gateway_info.Owner)
        sheet.cell(row=i, column=4, value=gateway_info.Brand)
        sheet.cell(row=i, column=5, value=gateway_info.Model)
        sheet.cell(row=i, column=6, value=gateway_info.Antenna_Type)
        sheet.cell(row=i, column=7, value=gateway_info.Antenna_DBI)
        sheet.cell(row=i, column=8, value=gateway_info.Firmware_Version)
        sheet.cell(row=i, column=9, value=gateway_info.VPN_KEY)
        sheet.cell(row=i, column=10, value=gateway_info.SSH_USER)
        sheet.cell(row=i, column=11, value=gateway_info.SSH_PASSWORD)
        sheet.cell(row=i, column=12, value=gateway_info.Default_USER)
        sheet.cell(row=i, column=13, value=gateway_info.Default_PASSWORD)
        sheet.cell(row=i, column=14, value=gateway_info.SERIAL)
        sheet.cell(row=i, column=15, value=gateway_info.Gateway_Token)  # Assuming TOKEN is a property of GatewayInfo

    # Save to file
    workbook.save(_finished_data_file_path)

def Gstack_setup(in_gateway_info: GatewayInfo, driver, sleep_time):
    driver = goto_Gateway_Page(driver, sleep_time)
    
    driver = goto_Add_Gateway_Page(driver)
    driver = setup_Organization(driver, _project.organization, sleep_time)
    driver = setup_Project_Department(driver, in_gateway_info.Project_Code, sleep_time)
    driver = setup_Gateway_Info(driver, in_gateway_info, sleep_time)
    time.sleep(10)
    driver, in_gateway_info = get_Gateway_Token(driver, in_gateway_info, sleep_time)
    time.sleep(sleep_time)
    return driver, in_gateway_info

def Revise_Model_To_UG67(in_EUI_list, driver, sleep_time):
    
    driver = goto_Gateway_Page(driver, sleep_time)
    driver = goto_hardware_page(driver, sleep_time)

    for in_EUI in in_EUI_list:
        start_time = time.time()
        driver = revise_Gateway_to_UG67(driver, in_EUI, sleep_time)
        print(in_EUI + " is revised to UG67")
        time.sleep(5)
        end_time = time.time()
        print(f"Total time to revise the record: in_EUI is {str(end_time - start_time)} seconds")
    return driver

_raw_data_file_path = 'Gateway_INFO.xlsx' # <-- Users need to revise the file path
#check file exit or not
if not os.path.exists(_raw_data_file_path):
    print("File not found")

_excel_page_name = 'DATA'

_project = EMSD_Project()
_gstack = GStack_Webpage()
sleep_time = 6
chrome_options = webdriver.ChromeOptions()
driver = webdriver.Chrome()
driver = login(driver, _gstack.Domain, _gstack.User_Name, _gstack.Password)
time.sleep(2)

gatewayInfos = read_Gateway_info_from_Excel(_raw_data_file_path, _excel_page_name)

for gatewayInfo in gatewayInfos:
    # if gatewayInfo.Gateway_Token is null or empty
    if not (gatewayInfo.Gateway_Token and gatewayInfo.Gateway_Token.strip()):
        #create a time to caculate the total time
        print(gatewayInfo.Brand)
        print(gatewayInfo.Model)
        print(gatewayInfo.EUI)
        print()

        start_time = time.time()
        driver, gatewayInfo = Gstack_setup(gatewayInfo, driver, sleep_time)
        save_gateway_info_to_excel(gatewayInfos, _raw_data_file_path, _excel_page_name)
        time.sleep(8)
        #close the new window
        driver.close()
        # focus on the main window
        driver.switch_to.window(driver.window_handles[0])
        end_time = time.time()
        # print out the total time to add the record
        print(f"Total time to add the record: {gatewayInfo.VPN_KEY}, {gatewayInfo.Brand}, {gatewayInfo.Model} is {str(end_time - start_time)} seconds")
        print("---------------------------------------------------------------------------------------------------------------------")    
        print()










