import openpyxl
import Common_Obj as co
from typing import List

class Excel_Controller:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def read_excel_and_convert_to_gateway_info(self, page_name) -> List[co.SIM_Info]:
        workbook = openpyxl.load_workbook(self.file_path)
        result_sheet = workbook[page_name]
        gateway_info_list = []

        for row in result_sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            gateway_info = co.SIM_Info(*row)
            gateway_info_list.append(gateway_info)

        return gateway_info_list

    # data dependency, need to improve in the future
    def save_gateway_info_to_excel(gateway_info_list: List[co.GatewayInfo], _finished_data_file_path: str, _sheet_name, headers: List[str] = None):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = _sheet_name
        
        for i, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=i, value=header)

        # Write data
        for i, gateway_info in enumerate(gateway_info_list, start=2):
            sheet.cell(row=i, column=1, value=gateway_info.EUI)
            sheet.cell(row=i, column=2, value=gateway_info.Project_Code)
            sheet.cell(row=i, column=3, value=gateway_info.Owner)
            sheet.cell(row=i, column=4, value=gateway_info.Brand)
            sheet.cell(row=i, column=5, value=gateway_info.Model)
            sheet.cell(row=i, column=6, value=gateway_info.Antenna_Type)
            sheet.cell(row=i, column=7, value=gateway_info.Antenna_DBI)
            sheet.cell(row=i, column=8, value=gateway_info.Firmware_Version)
            sheet.cell(row=i, column=9, value=gateway_info.VPN_KEY)
            sheet.cell(row=i, column=10, value=gateway_info.SSH_USER)
            sheet.cell(row=i, column=11, value=gateway_info.SSH_PASSWORD)
            sheet.cell(row=i, column=12, value=gateway_info.Default_USER)
            sheet.cell(row=i, column=13, value=gateway_info.Default_PASSWORD)
            sheet.cell(row=i, column=14, value=gateway_info.SERIAL)
            sheet.cell(row=i, column=15, value=gateway_info.Gateway_Token)  # Assuming TOKEN is a property of GatewayInfo

        # Save to file
        workbook.save(_finished_data_file_path)
